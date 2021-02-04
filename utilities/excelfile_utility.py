import openpyxl

#reusable method which will return row count
def getrowcount(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return(sheet.max_row)

# It will return the coln count
def colcount(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return(sheet.max_column)

# It will read the cell of the sheet
def readrow(filepath,sheetname,r_num,c_num):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return sheet.cell(row=r_num,column=c_num).value

